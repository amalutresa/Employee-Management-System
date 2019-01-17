from tkinter import *
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox
import pymysql
import pymysql.cursors
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range

w_front=w_table=w_enter=w=0

con=pymysql.connect('localhost','root','','employee')
a=con.cursor()

def Front():
    global w_front
    w_front=Tk()
    w_front.title("Employee Management System")
    Label(w_front,text="----- CHOOSE A OPTION -----",fg='blue',width=40).grid(row=2,columnspan=5)

    b_TALLY=Button(w_front,text="Display Table!",command=Tally,width=40).grid(row=4,columnspan=8,rowspan=2)
    b_ADD=Button(w_front,text="Add Record",command=Enter,width=30).grid(row=6,columnspan=8,rowspan=2)
    b_DELETE=Button(w_front,text="Delete Record",command=Delete_Front,width=30).grid(row=8,columnspan=8,rowspan=2)
    b_SEARCH=Button(w_front,text="Search Record",command=Search_Front,width=30).grid(row=10,columnspan=8,rowspan=2)
    b_SAVE=Button(w_front,text="Save to xlsx",command=Save,width=30).grid(row=12,columnspan=8,rowspan=2)
    b_ABOUT=Button(w_front,text="About",command=About,width=30).grid(row=14,columnspan=8,rowspan=2)
    b_EXIT=Button(w_front,text="Exit",command=Exit,width=40).grid(row=16,columnspan=8,rowspan=2)

def Enter():
    global w_enter,eID,eFName,eLName,eAge,eIncome,sSex,eContact,sDes,sCity
    w_enter=Tk()
    w_enter.title("ENTER THE DETAILS OF THE NEW EMPLOYEE")
    Label(w_enter,text="Emp ID",fg='red').grid(row=2,columnspan=3)
    Label(w_enter,text="First Name").grid(row=3,columnspan=3)
    Label(w_enter,text="Last Name").grid(row=4,columnspan=3)
    Label(w_enter,text="Age").grid(row=5,columnspan=3)
    Label(w_enter,text="Income").grid(row=6,columnspan=3)
    Label(w_enter,text="Contact").grid(row=7,columnspan=3)
    Label(w_enter,text="Sex").grid(row=9,columnspan=3)
    Label(w_enter,text="Designation").grid(row=10,columnspan=3)
    Label(w_enter,text="Opr City").grid(row=11,columnspan=3)

    Label(w_enter,text=":").grid(row=2,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=3,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=4,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=5,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=6,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=7,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=9,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=10,column=2,columnspan=2)
    Label(w_enter,text=":").grid(row=11,column=2,columnspan=2)

    Label(w_enter,text="              ").grid(row=8,columnspan=10)

    eID=Entry(w_enter,width=22)
    eID.grid(row=2,column=3,columnspan=4)
    eFName=Entry(w_enter,width=22)
    eFName.grid(row=3,column=3,columnspan=4)
    eLName=Entry(w_enter,width=22)
    eLName.grid(row=4,column=3,columnspan=4)
    eAge=Entry(w_enter,width=22)
    eAge.grid(row=5,column=3,columnspan=4)
    eIncome=Entry(w_enter,width=22)
    eIncome.grid(row=6,column=3,columnspan=4)
    eContact=Entry(w_enter,width=22)
    eContact.grid(row=7,column=3,columnspan=4)
    
    sSex=Spinbox(w_enter,values=('M','F'))
    sSex.grid(row=9,column=3,columnspan=4)
    sDes=Spinbox(w_enter,values=('CEO','PROJECT HEAD','TEAM LEADER','ANALYST','DEVELOPER','ACCOUNTANT','OTHER'))
    sDes.grid(row=10,column=3,columnspan=4)
    sCity=Spinbox(w_enter,values=('SAN FRANSISCO','WINDSOR','NOIDA','BENGLURU'))
    sCity.grid(row=11,column=3,columnspan=4)
    
    Label(w_enter,text="              ").grid(row=12,columnspan=10)
    
    eID.focus()

    b_ENTER=Button(w_enter,text="Add Record",command=Add,width=20).grid(row=13,columnspan=3,rowspan=2)
    b_UPDATE=Button(w_enter,text="Update Record",command=Update,width=20).grid(row=13,column=3,columnspan=3,rowspan=2)
    b_CLEAR=Button(w_enter,text="Clear Fields",command=Clear,width=20).grid(row=15,columnspan=3,rowspan=2)
    b_EXIT=Button(w_enter,text="Exit",command=Exit2,width=20).grid(row=15,column=3,columnspan=3,rowspan=2)
    
def Tally():
    global w_table
    w_table=Tk()
    w_table.title("Database: ")
    st=ScrolledText(w_table,width=100)
    st.pack(padx=10,pady=10)
    rp=("Select * from personal")
    a.execute(rp)
    rep=a.fetchall()
    st.insert(INSERT,"\t\t\t\t\tPERSONAL INFORMATION : \n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t---\t\t---\t\t----------\n")
    st.insert(INSERT,"EMP_ID\t\tFIRST_NAME\t\tLAST_NAME\t\tAGE\t\tSEX\t\t CONTACT\n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t---\t\t---\t\t----------\n")
    for row in rep:
        ei=row[0]
        fn=row[1]
        ln=row[2]
        age=row[3]
        sex=row[4]
        cont=row[5]
        st.insert(INSERT,("%s\t\t%s\t\t%s\t\t%d\t\t%c\t\t%d\n"%(ei,fn,ln,age,sex,cont)))

    rc=("Select * from corporate")
    a.execute(rc)
    rec=a.fetchall()
    st.insert(INSERT,"\n\n\n\t\t\t\t\tCORPORATE INFORMATION : \n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t-----------\t\t--------\t\t------\n")
    st.insert(INSERT,"EMP_ID\t\tFIRST_NAME\t\tLAST_NAME\t\tDESIGNATION\t\tOPR_CITY\t\tINCOME\n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t-----------\t\t--------\t\t------\n")
    for row in rec:
        ei=row[0]
        fn=row[1]
        ln=row[2]
        des=row[3]
        city=row[4]
        inc=row[5]
        st.insert(INSERT,("%s\t\t%s\t\t%s\t\t%s\t\t%s\t\t%d\n"%(ei,fn,ln,des,city,inc)))

    
    con.commit()

def Save():
    wb = Workbook()
    #wb = load_workbook(filename = 'EMS.xlsx')
    
    rp=("Select * from personal")
    a.execute(rp)
    rep=a.fetchall()
    wsp = wb.active
    wsp.title = "Personal data"
    wsp.append(["EMP_ID","FIRST_NAME","LAST_NAME","AGE","SEX","CONTACT"])
    for row in rep:
        wsp.append(row)
    
    rc=("Select * from corporate")
    a.execute(rc)
    rec=a.fetchall()
    wsc = wb.create_sheet(title="Corporate Data")
    #wsc = wb.active
    wsc.append(["EMP_ID","FIRST_NAME","LAST_NAME","DESIGNATION","OPR_CITY","INCOME"])
    for row in rec:
        wsc.append(row)
        
    print("Exported data to EMS.xlsx")        
     
    wb.save("EMS.xlsx")



    
def Add():    
    ei=eID.get()
    fn=eFName.get()
    ln=eLName.get()
    age=int(eAge.get())
    sex=sSex.get()
    inc=int(eIncome.get())
    cont=float(eContact.get())
    des=sDes.get()
    city=sCity.get()
    
    add_p="Insert into personal values('%s','%s','%s',%d,'%c',%f)"%(ei,fn,ln,age,sex,cont)
    a.execute(add_p)
    print("Record Added Successfully in personal table!\n...")
    
    add_c="Insert into corporate values('%s','%s','%s','%s','%s',%d)"%(ei,fn,ln,des,city,inc)
    a.execute(add_c)
    print("Record Added Successfully in corporate table!\n")

    con.commit()
    Clear()

def Clear():
    eID.delete(0,END)
    eFName.delete(0,END)
    eLName.delete(0,END)
    eAge.delete(0,END)
    eIncome.delete(0,END)
    sSex.delete(0,END)
    eContact.delete(0,END)
    sDes.delete(0,END)
    sCity.delete(0,END)

    eID.focus()

def Delete_Front():
    global w_dfront,eIDd
    w_dfront=Tk()
    w_dfront.title("DELETE A RECORD")
    
    Label(w_dfront,text="Emp ID",fg='red').grid(row=2,columnspan=3)
    Label(w_dfront,text=":").grid(row=2,column=2,columnspan=2)
    eIDd=Entry(w_dfront,width=22)
    eIDd.grid(row=2,column=3,columnspan=3)

    Label(w_dfront,text="              ").grid(row=3,columnspan=8)
    
    b_SEARCH=Button(w_dfront,text="DELETE",command=Delete,width=20).grid(row=4,columnspan=3)
    b_EXIT=Button(w_dfront,text="EXIT",command=Exit3,width=20).grid(row=4,column=3,columnspan=3)
    
def Delete():
    eids=eIDd.get()
    dc="Delete from corporate where EMP_ID='%s'"%(eids)
    a.execute(dc)
    
    dp="Delete from personal where EMP_ID='%s'"%(eids)
    a.execute(dp)
    
    print("Record Deleted Successfully!")
    con.commit()

def Update():
    ei=eID.get()
    fn=eFName.get()
    ln=eLName.get()
    age=int(eAge.get())
    sex=sSex.get()
    inc=int(eIncome.get())
    cont=float(eContact.get())
    des=sDes.get()
    city=sCity.get()
    
    up="Update personal set  First_name='%s', Last_name='%s', Age='%d', Sex='%c', Contact='%d' where Emp_ID='%s'"%(fn,ln,age,sex,cont,ei)
    a.execute(up)
    print("Record Updated Successfully in personal table")

    uc="Update corporate set  First_name='%s', Last_name='%s', Designation='%s', OPR_City='%s', Income='%d' where Emp_ID='%s'"%(fn,ln,des,city,inc,ei)
    a.execute(uc)
    print("Record Updated Successfully in corporate table")

    con.commit()

def Search_Front():
    global w_sfront,eIDs
    w_sfront=Tk()
    w_sfront.title("SEARCH RECORD")
    
    Label(w_sfront,text="Emp ID",fg='red').grid(row=2,columnspan=3)
    Label(w_sfront,text=":").grid(row=2,column=2,columnspan=2)
    eIDs=Entry(w_sfront,width=22)
    eIDs.grid(row=2,column=3,columnspan=3)

    Label(w_enter,text="              ").grid(row=3,columnspan=8)
    
    b_SEARCH=Button(w_sfront,text="SEARCH",command=Search,width=20).grid(row=4,columnspan=3)
    b_EXIT=Button(w_sfront,text="EXIT",command=Exit4,width=20).grid(row=4,column=3,columnspan=3)
    
def Search():
    w_table=Tk()
    st=ScrolledText(w_table,width=100)
    st.pack(padx=10,pady=10)
    eids=eIDs.get()
    
    sep="Select * from personal where EMP_ID='%s'"%(eids)
    a.execute(sep)
    rp=a.fetchall()
    st.insert(INSERT,"\t\t\t\t\tPERSONAL INFORMATION : \n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t---\t\t---\t\t----------\n")
    st.insert(INSERT,"EMP_ID\t\tFIRST_NAME\t\tLAST_NAME\t\tAGE\t\tSEX\t\t CONTACT\n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t---\t\t---\t\t----------\n")
    for row in rp:
        ei=row[0]
        fn=row[1]
        ln=row[2]
        age=row[3]
        sex=row[4]
        cont=row[5]
        st.insert(INSERT,("%s\t\t%s\t\t%s\t\t%d\t\t%c\t\t%d\n"%(ei,fn,ln,age,sex,cont)))

    sec="Select * from corporate where EMP_ID='%s'"%(eids)
    a.execute(sec)
    rc=a.fetchall()
    st.insert(INSERT,"\n\n\n\t\t\t\t\tCORPORATE INFORMATION : \n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t-----------\t\t--------\t\t------\n")
    st.insert(INSERT,"EMP_ID\t\tFIRST_NAME\t\tLAST_NAME\t\tDESIGNATION\t\tOPR_CITY\t\tINCOME\n")
    st.insert(INSERT,"------\t\t----------\t\t---------\t\t-----------\t\t--------\t\t------\n")
    for row in rc:
        ei=row[0]
        fn=row[1]
        ln=row[2]
        des=row[3]
        city=row[4]
        inc=row[5]
        st.insert(INSERT,("%s\t\t%s\t\t%s\t\t%s\t\t%s\t\t%d\n"%(ei,fn,ln,des,city,inc)))

    
    con.commit()

def About():
    w_table=Tk()
    st=ScrolledText(w_table,width=80)
    st.pack(padx=10,pady=10)
    st.insert(INSERT,"\t\t\tABOUT EMPLOYEE MANAGEMENT SYSTEM\n\n\tThis project is programmed and designed by Surya Bansal\n\n\tAs a major project to be submitted at RIMT-IET\n\n\t\t* B.tech - CSE\n\t\t  7th-A\n\t\t  1420147")
    
def Check():
    l=ee1.get()
    p=ee2.get()
    if l=="admin" and p=="password1":
        Front()
        w.destroy()
    else:
        messagebox.showwarning("Wrong Login!","Enter Again")
        ee1.delete(0,END)
        ee2.delete(0,END)

def Exit1():
    w.destroy()

def Exit2():
    w_enter.destroy()

def Exit3():
    w_dfront.destroy()
    
def Exit4():
    w_sfront.destroy()

def Exit():
    a.close()
    con.close()
    w_front.destroy()
    w_table.destroy()
    sys.exit()

w=Tk()
w.title("EMS")
Label(w,text="WELCOME TO EMPLOYEE MANAGEMENT SYSTEM",bg='#00ccff').grid(row=1, columnspan=4, sticky=N)
Label(w,text="Enter User Name:").grid(row=3,column=1)
ee1=Entry(w,bd=3,fg='red')
ee1.grid(row=3,column=3)
Label(w,text="Enter Password:").grid(row=4,column=1)
ee2=Entry(w,bd=3,show="*",fg='red')
ee2.grid(row=4,column=3)
Button(w,text="Sign In",command=Check).grid(row=6,column=1)
Button(w,text="Exit",command=Exit1).grid(row=6,column=3)
ee1.focus()

w.mainloop()
